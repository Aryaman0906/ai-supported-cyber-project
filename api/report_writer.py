"""Daily Markdown/CSV/XLSX report generation for Gmail scan results."""
from __future__ import annotations

from collections import Counter
import csv, html, importlib.util, io, zipfile
from typing import Any


def summarize_results(results: list[dict[str, Any]]) -> dict[str, Any]:
    counts = Counter(result.get("risk_level", "unknown") for result in results)
    suspicious = [r for r in results if r.get("risk_level") in {"high", "medium", "unknown"}]
    return {
        "total_scanned": len(results), "high_count": counts["high"], "medium_count": counts["medium"], "low_count": counts["low"], "unknown_count": counts["unknown"],
        "top_sender_domains": Counter(r.get("sender_domain", "") for r in suspicious if r.get("sender_domain")).most_common(10),
        "top_url_domains": Counter(domain for r in suspicious for domain in r.get("url_domains", [])).most_common(10),
        "high_risk_messages": [r for r in results if r.get("risk_level") == "high"][:25],
    }


def generate_markdown_report(date: str, results: list[dict[str, Any]]) -> str:
    summary = summarize_results(results)
    lines = [f"# AI Cyber Daily Gmail Report - {date}", "", "## Summary", "", f"- Total scanned: {summary['total_scanned']}", f"- High risk: {summary['high_count']}", f"- Medium risk: {summary['medium_count']}", f"- Low risk: {summary['low_count']}", f"- Unknown: {summary['unknown_count']}", "", "## Top suspicious sender domains"]
    lines += [f"- {d}: {c}" for d, c in summary["top_sender_domains"]] or ["- None"]
    lines += ["", "## Top suspicious URL domains"]
    lines += [f"- {d}: {c}" for d, c in summary["top_url_domains"]] or ["- None"]
    lines += ["", "## High risk message summaries"]
    for r in summary["high_risk_messages"]:
        lines.append(f"- Message `{r.get('message_id')}` from `{r.get('sender_domain')}` score={r.get('score')} reasons={'; '.join(r.get('reasons', [])[:3])}")
    if not summary["high_risk_messages"]: lines.append("- None")
    lines += ["", "## Safety note", "This report is an assistive defensive signal only. Do not open suspicious links or attachments. Human review is required."]
    return "\n".join(lines) + "\n"


def generate_csv_report(results: list[dict[str, Any]]) -> str:
    output = io.StringIO()
    fields = ["message_id", "date", "sender_domain", "risk_level", "score", "url_domains", "reasons"]
    writer = csv.DictWriter(output, fieldnames=fields); writer.writeheader()
    for r in results:
        writer.writerow({"message_id": r.get("message_id", ""), "date": r.get("date", ""), "sender_domain": r.get("sender_domain", ""), "risk_level": r.get("risk_level", ""), "score": r.get("score", ""), "url_domains": ";".join(r.get("url_domains", [])), "reasons": " | ".join(r.get("reasons", []))})
    return output.getvalue()



def _worksheet_xml(rows: list[list[Any]]) -> str:
    row_xml: list[str] = []
    for row_number, row in enumerate(rows, start=1):
        cells = []
        for column_number, value in enumerate(row, start=1):
            column = chr(ord("A") + column_number - 1)
            safe_value = html.escape(str(value or ""))
            cells.append(f'<c r="{column}{row_number}" t="inlineStr"><is><t>{safe_value}</t></is></c>')
        row_xml.append(f'<row r="{row_number}">{"".join(cells)}</row>')
    return "".join(row_xml)


def _generate_basic_xlsx_report_bytes(rows: list[list[Any]]) -> bytes:
    worksheet = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f'<sheetData>{_worksheet_xml(rows)}</sheetData></worksheet>'
    )
    workbook = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        '<sheets><sheet name="Gmail Report" sheetId="1" r:id="rId1"/></sheets></workbook>'
    )
    rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" '
        'Target="xl/workbook.xml"/></Relationships>'
    )
    workbook_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" '
        'Target="worksheets/sheet1.xml"/></Relationships>'
    )
    content_types = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        '<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        '</Types>'
    )
    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as workbook_zip:
        workbook_zip.writestr("[Content_Types].xml", content_types)
        workbook_zip.writestr("_rels/.rels", rels)
        workbook_zip.writestr("xl/workbook.xml", workbook)
        workbook_zip.writestr("xl/_rels/workbook.xml.rels", workbook_rels)
        workbook_zip.writestr("xl/worksheets/sheet1.xml", worksheet)
    return output.getvalue()


def _clean_xlsx_value(value: Any) -> str:
    """Normalize values so they are safe and readable in Excel cells."""
    if value is None:
        return ""
    if isinstance(value, (list, tuple, set)):
        value = "; ".join(_clean_xlsx_value(item) for item in value)
    text = _illegal_characters_re().sub("", str(value))
    return " ".join(text.replace("\r", " ").replace("\n", " ").split())


def _first_present_value(record: dict[str, Any], keys: tuple[str, ...]) -> Any:
    for key in keys:
        value = record.get(key)
        if value is not None and value != "":
            return value
    return ""


def _illegal_characters_re():
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

    return ILLEGAL_CHARACTERS_RE


def _risk_fill(risk_level: str):
    normalized_risk = risk_level.strip().lower()
    from openpyxl.styles import PatternFill

    if normalized_risk in {"high", "phishing", "dangerous", "malicious"}:
        return PatternFill(fill_type="solid", fgColor="FCE4D6")
    if normalized_risk in {"medium", "suspicious", "warning"}:
        return PatternFill(fill_type="solid", fgColor="FFF2CC")
    if normalized_risk in {"low", "safe", "legitimate", "clean"}:
        return PatternFill(fill_type="solid", fgColor="E2F0D9")
    return PatternFill(fill_type="solid", fgColor="E7E6E6")


def generate_xlsx_report_bytes(results: list[dict[str, Any]]) -> bytes:
    """Generate a formatted XLSX workbook for local Gmail scan results."""
    columns: list[tuple[str, tuple[str, ...], int]] = [
        ("Message ID", ("message_id", "id", "gmail_message_id"), 28),
        ("Scanned At", ("created_at", "date", "scanned_at", "scan_time", "timestamp"), 22),
        ("Sender", ("sender", "from", "from_email"), 32),
        ("Sender Domain", ("sender_domain",), 24),
        ("Subject", ("subject_preview", "subject", "email_subject"), 42),
        ("Risk Level", ("risk_level", "risk", "prediction"), 16),
        ("Score", ("score", "risk_score", "confidence"), 12),
        ("Gmail Label", ("labels_applied", "label", "gmail_label", "applied_label"), 24),
        ("Reason", ("reasons", "reason", "explanation", "summary", "details"), 60),
    ]
    headers = [header for header, _, _ in columns]

    if importlib.util.find_spec("openpyxl") is None:
        rows = [headers]
        for result in results:
            rows.append([str(_first_present_value(result, keys) or "") for _, keys, _ in columns])
        return _generate_basic_xlsx_report_bytes(rows)

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Gmail Report"

    worksheet.append(headers)

    for result in results:
        worksheet.append([_clean_xlsx_value(_first_present_value(result, keys)) for _, keys, _ in columns])

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin_side = Side(style="thin", color="D9E2F3")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        fill = _risk_fill(str(row[5].value or ""))
        for cell in row:
            cell.fill = fill
            cell.alignment = wrap_alignment
            cell.border = border

    for column_number, (_, _, width) in enumerate(columns, start=1):
        worksheet.column_dimensions[get_column_letter(column_number)].width = width

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    table_ref = f"A1:{get_column_letter(len(columns))}{max(worksheet.max_row, 1)}"
    table = Table(displayName="GmailPollReport", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    worksheet.add_table(table)

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()
